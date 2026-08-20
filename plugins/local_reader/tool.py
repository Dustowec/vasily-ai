"""Local Reader plugin - reads files from allowed directories.

Supports: csv, json, txt, md, xlsx (via openpyxl), pdf (via PyPDF2).
Protected against path traversal attacks.
"""

import csv
import json
import os
from pathlib import Path
from typing import Any

from core.base_tool import BaseTool
from core.plugin_types import make_error

# Разрешённые директории (относительно корня проекта)
ALLOWED_DIRS = ["data", "reports"]

# Поддерживаемые расширения и их парсеры
PARSERS = {
    ".csv": "csv",
    ".json": "json",
    ".txt": "text",
    ".md": "text",
    ".xlsx": "xlsx",
    ".pdf": "pdf",
}

MAX_FILE_SIZE_BYTES = 1024 * 1024  # 1 MB


class LocalReaderTool(BaseTool):
    """Read local files from allowed directories."""

    name = "local_reader"
    description = "Read contents of files from data/ and reports/ directories"
    version = "1.1.0"

    async def _execute(self, path: str = "", **kwargs) -> dict[str, Any]:
        """Read a file from an allowed directory."""
        if not path:
            return make_error(
                "invalid_url",
                "File path is required",
                "Provide a path to a file in data/ or reports/ directory.",
            )

        # Проверяем path traversal
        if not self._is_path_allowed(path):
            return make_error(
                "invalid_url",
                f"Path '{path}' is not allowed. Only files in {ALLOWED_DIRS} are accessible.",
                "Use a path within data/ or reports/ directories.",
            )

        file_path = Path(path)
        if not file_path.exists():
            return make_error(
                "invalid_url",
                f"File not found: {path}",
                "Check that the file exists and the path is correct.",
            )

        if not file_path.is_file():
            return make_error(
                "invalid_url",
                f"Path is not a file: {path}",
                "Provide a path to a file, not a directory.",
            )

        # Проверяем размер
        size = file_path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            return make_error(
                "invalid_url",
                f"File too large: {size} bytes (max {MAX_FILE_SIZE_BYTES})",
                "Provide a smaller file or increase MAX_FILE_SIZE_BYTES.",
            )

        # Определяем парсер по расширению
        ext = file_path.suffix.lower()
        parser_type = PARSERS.get(ext)

        if parser_type is None:
            return make_error(
                "invalid_url",
                f"Unsupported file type: {ext}",
                f"Supported types: {', '.join(PARSERS.keys())}",
            )

        try:
            content = self._parse_file(file_path, parser_type)
            return {
                "status": "success",
                "source": "local",
                "path": str(file_path),
                "size_bytes": size,
                "parser": parser_type,
                "content": content,
            }
        except Exception as e:
            return make_error(
                "connection_failed",
                f"Failed to read file: {e}",
                "Check that the file is not corrupted and has the correct format.",
            )

    def _is_path_allowed(self, path: str) -> bool:
        """Check if path is within allowed directories."""
        # Получаем абсолютный путь
        abs_path = os.path.abspath(path)
        cwd = os.getcwd()

        # Проверяем каждую разрешённую директорию
        for allowed in ALLOWED_DIRS:
            # Разрешённый путь относительно CWD
            allowed_path = os.path.abspath(os.path.join(cwd, allowed))
            # Проверяем, что abs_path начинается с allowed_path
            if abs_path.startswith(allowed_path):
                # Дополнительная проверка: убеждаемся, что не вышли за пределы
                # через .. (например, data/../../etc/passwd)
                normalized = os.path.normpath(abs_path)
                if normalized.startswith(allowed_path):
                    return True
        return False

    def _parse_file(self, file_path: Path, parser_type: str) -> Any:
        """Parse file content based on parser type."""
        if parser_type == "text":
            return file_path.read_text(encoding="utf-8", errors="replace")

        if parser_type == "json":
            with open(file_path, encoding="utf-8") as f:
                return json.load(f)

        if parser_type == "csv":
            with open(file_path, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                return list(reader)

        if parser_type == "xlsx":
            try:
                import openpyxl
            except ImportError:
                return {"error": "openpyxl is not installed. Run: pip install openpyxl"}

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            result = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(list(row))
                result[sheet_name] = rows
            wb.close()
            return result

        if parser_type == "pdf":
            try:
                import PyPDF2
            except ImportError:
                return {"error": "PyPDF2 is not installed. Run: pip install PyPDF2"}

            try:
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in reader.pages:
                        extracted = page.extract_text()
                        if extracted:
                            text += extracted + "\n"
                    return text.strip() or "[No text extracted from PDF]"
            except Exception as e:
                return {"error": f"Failed to read PDF: {e}"}

        return {"error": f"Unknown parser: {parser_type}"}

    def _get_parameters(self) -> dict[str, Any]:
        """Get parameter schema."""
        return {
            "path": {
                "type": "string",
                "description": f"Path to file in {ALLOWED_DIRS} directories",
                "required": True,
            },
        }
